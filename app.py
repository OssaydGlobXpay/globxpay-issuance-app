import streamlit as st
import pandas as pd
import csv
from datetime import datetime
from io import BytesIO, StringIO
from openpyxl import load_workbook
from PIL import Image


# === PAGE CONFIG ===
st.set_page_config(page_title="GlobXpay Issuance Tool", layout="wide")

logo = Image.open("GlobXpay logo wa-01.jpg")
st.sidebar.image(logo, width=160)
st.sidebar.title("GlobXpay Issuance Center")

tab1, tab2 = st.tabs(["📥 Convert CMS Template to CSC CSV", "📆 Daily Issuance to Bulk Template"])

# === Mandatory field config ===
mandatory_fields = {
    0: "Record Date", 1: "Institution Number", 2: "Branch", 3: "Application Number", 4: "Action",
    7: "Customer ID", 9: "Entity Type", 11: "First Name", 13: "Last Name", 14: "Birth Date",
    16: "Identity Type", 17: "Identity Number", 19: "Gender", 20: "Nationality", 21: "Phone Number",
    24: "City Code", 25: "Country Code", 26: "Address Line 1", 32: "Bank Account Number",
    34: "Account Name", 35: "Credit Limit", 38: "Cardholder Name", 44: "Product Code", 53: "ID Expiry Date"
}

# === Convert CMS Excel to Clean CSV ===
def convert_to_csv(data):
    string_buffer = StringIO()
    writer = csv.writer(string_buffer, delimiter=";", quoting=csv.QUOTE_NONE, escapechar="\\", lineterminator="\n")
    writer.writerows(data)
    return string_buffer.getvalue().encode("utf-8")

# === Clean & Validate CMS Template ===
def validate_and_clean(df):
    errors = []
    cleaned_rows = []
    for idx, row in df.iterrows():
        row_errors = []
        row_values = []
        for i in range(76):
            val = str(row[i]).strip() if i < len(row) and pd.notna(row[i]) else ""
            if i in mandatory_fields and not val:
                row_errors.append(f"Missing '{mandatory_fields[i]}'")
            row_values.append(val)
        if row_errors:
            errors.append(f"🔴 Row {idx + 2}: " + ", ".join(row_errors))
        cleaned_rows.append(row_values)
    return cleaned_rows, errors

# === TAB 1: CMS to CSV ===
with tab1:
    st.markdown("## 📥 Convert CMS_Bulk_Cards_Full → CSC CSV")
    uploaded_file = st.file_uploader("📤 Upload CMS Excel File (.xlsx)", type=["xlsx"], key="cms_upload")

    if uploaded_file:
        try:
            df = pd.read_excel(uploaded_file, dtype=str).fillna("").astype(str)
            df = df.iloc[:, :76]
            st.success(f"✅ Loaded {len(df)} rows")
            cleaned_data, errors = validate_and_clean(df)

            if errors:
                st.error("⚠️ Issues Found – Fix Before Download")
                with st.expander("Click to expand details"):
                    for e in errors:
                        st.markdown(f"- {e}")
            else:
                st.success("✅ All data valid. Ready to download.")
                csv_data = convert_to_csv(cleaned_data)
                file_name = f"CSC_Converted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                st.download_button("📥 Download CSC CSV File", data=csv_data, file_name=file_name, mime="text/csv")

        except Exception as e:
            st.error(f"❌ Error reading file: {e}")

# === TAB 2: Daily to Bulk Template ===
with tab2:
    st.markdown("## 📆 Convert Daily Customer Extract to Bulk Template")
    uploaded_daily = st.file_uploader("📤 Upload Daily Extract (.xlsx)", type=["xlsx"], key="daily_upload")

    if uploaded_daily:
        try:
            # Load template
            template_path = "CMS_Bulk_Cards_Full.xlsx"
            wb = load_workbook(template_path)
            ws = wb.active

            # Load daily extract
            df = pd.read_excel(uploaded_daily).fillna("")

            # Product Code Lookup
            product_lookup = {1: "1201", 7: "1203"}

            start_row = 2  # Fill from row 2
            for i, row in df.iterrows():
                r = start_row + i
                fullname = row.get("Full Name in English", "").split()
                first, middle, last = (fullname + ["", "", ""])[:3]
                phone = str(row.get("Phone number", "")).strip()
                if not phone.startswith("00"):
                    phone = "00" + phone
                expiry = pd.to_datetime(row.get("Identification expiry date", ""), errors="coerce")
                expiry_str = expiry.strftime("%d/%m/%Y") if pd.notna(expiry) else ""
                card_type = int(row.get("Card Type", 0))
                product_code = product_lookup.get(card_type, "")

                # Fill mapped fields
                ws.cell(r, 1).value = datetime.today().strftime("%Y%m%d")
                ws.cell(r, 2).value = "012"
                ws.cell(r, 3).value = "0012"
                ws.cell(r, 4).value = f"6042025{str(i+1).zfill(4)}"
                ws.cell(r, 5).value = "C"
                ws.cell(r, 6).value = "C"
                ws.cell(r, 7).value = "N"
                ws.cell(r, 8).value = row.get("Identity number", "")
                ws.cell(r, 10).value = "37"
                ws.cell(r, 12).value = first
                ws.cell(r, 13).value = middle
                ws.cell(r, 14).value = last
                ws.cell(r, 17).value = "01"
                ws.cell(r, 18).value = row.get("Identity number", "")
                ws.cell(r, 19).value = "01"
                ws.cell(r, 21).value = "400"
                ws.cell(r, 22).value = phone
                ws.cell(r, 25).value = "001"
                ws.cell(r, 26).value = "400"
                ws.cell(r, 27).value = row.get("Street", "")
                ws.cell(r, 32).value = "2"
                ws.cell(r, 33).value = "N"
                ws.cell(r, 34).value = row.get("Bank account", "")
                ws.cell(r, 35).value = row.get("Full Name in English", "")
                ws.cell(r, 37).value = "0"  # Credit Limit always 0
                ws.cell(r, 39).value = row.get("Full Name in English", "")
                ws.cell(r, 45).value = product_code
                ws.cell(r, 54).value = expiry_str  # Column BB = 54

                for col in range(1, 77):
                    ws.cell(r, col).number_format = "@"

            # Save and offer download
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"CMS_Bulk_Cards_Full_{datetime.now().strftime('%Y%m%d')}.xlsx"
            st.success("✅ CMS Bulk Template filled successfully.")
            st.download_button("📥 Download CMS_Bulk_Cards_Full.xlsx", data=output, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        except Exception as e:
            st.error(f"❌ Failed to generate bulk file:\n{e}")
